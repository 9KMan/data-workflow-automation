"""Excel/CSV parser service for data ingestion."""
import io
from typing import Iterator

from openpyxl import load_workbook


def parse_excel(content: bytes) -> Iterator[dict]:
    """
    Parse .xlsx file from bytes and yield rows as dicts.
    Processes all sheets; each sheet's rows are prefixed with _sheet key.
    """
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = None
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [str(h).strip() if h else f"column_{i}" for i, h in enumerate(row)]
                continue
            if all(v is None for v in row):
                continue  # skip blank rows
            row_dict = dict(zip(headers, row))
            row_dict["_sheet"] = sheet_name
            yield row_dict
    wb.close()


def count_excel_rows(content: bytes) -> int:
    """Count total data rows across all sheets (excluding headers)."""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    count = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_passed = False
        for row in ws.iter_rows(values_only=True):
            if not header_passed:
                header_passed = True
                continue
            if any(v is not None for v in row):
                count += 1
    wb.close()
    return count